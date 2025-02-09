from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse
import pdfplumber
import openpyxl
from openpyxl.styles import Font
import os
from fastapi.middleware.cors import CORSMiddleware
import re
from datetime import datetime

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://medical-pdf-to-excel-1.onrender.com/"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)

def extract_date_from_filename(filename):
    # Регулярные выражения для разных форматов даты
    patterns = [
        r'(\d{4})(\d{2})(\d{2})',  # YYYYMMDD
        r'(\d{2})\.(\d{2})\.(\d{4})',  # DD.MM.YYYY
        r'(\d{2})\s(\d{2})\s(\d{4})',  # DD MM YYYY
        r'(\d{4})\s(\d{2})\s(\d{2})'   # YYYY MM DD
    ]
    
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            groups = match.groups()
            if len(groups[0]) == 4:  # YYYYMMDD или YYYY MM DD
                return datetime(int(groups[0]), int(groups[1]), int(groups[2]))
            else:  # DD.MM.YYYY или DD MM YYYY
                return datetime(int(groups[2]), int(groups[1]), int(groups[0]))
    
    raise ValueError(f"Не удалось извлечь дату из имени файла: {filename}")

def parse_line(line):
    irrelevant_patterns = [
        r"Пол:", r"Возраст:", r"ИНЗ:", r"Дата взятия", r"Дата поступления", r"Врач:", r"Дата печати",
        r"стр\.\s?\d+\sиз\s\d+", r"Исследуемый материал", r"Хранение и транспортировка"
    ]
    if any(re.search(pattern, line, re.IGNORECASE) for pattern in irrelevant_patterns):
        return None

    pattern = r"(.+?)\s+([\d.,↓↑+-]+)\s+([\w/%^°μгл]+)?\s+([\d.,-]+)?"
    match = re.match(pattern, line)
    if match:
        return tuple(cell if cell is not None else "" for cell in match.groups())
    return None

def extract_data_from_pdf(file_path):
    try:
        with pdfplumber.open(file_path) as pdf:
            table_data = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines = text.split("\n")
                    for line in lines:
                        parsed_line = parse_line(line)
                        if parsed_line and all(parsed_line):
                            table_data.append(parsed_line)
            return table_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке PDF: {str(e)}")

def is_relevant_row(row):
    if not row or all(cell.strip() == "" for cell in row):
        return False

    irrelevant_patterns = [
        r"Гематология",
        r"Биохимические исследования",
        r"Наименование теста.*Результат.*Единицы.*Референсные",
        r"Дата выдачи:",
        r"Стр\\стр\.\?\s+\d+\s+из\s+\d+",
        r"врач КЛД",
        r"<\>\\=",
        r"от\s+\d+\s+(лет|месяцев|недель|дней)",
        r"\d+\s+(лет|месяцев|недель|дней)",
    ]

    if any(re.search(pattern, " ".join(row), re.IGNORECASE) for pattern in irrelevant_patterns):
        return False

    if not any(re.search(r"\d", cell) for cell in row):
        return False

    if not re.search(r"[\d.,]", row[1]):
        return False

    forbidden_chars_pattern = r"[><=/'\"\d]"
    if re.search(forbidden_chars_pattern, row[0]):
        return False

    return True

def create_excel_from_data(data_by_date, output_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Анализы"
    
    # Начальная колонка для первой даты
    current_col = 2
    
    # Записываем названия исследований в первую колонку
    all_tests = sorted(set(test for date_data in data_by_date.values() 
                          for test, _, _, _ in date_data))
    
    for row, test in enumerate(all_tests, start=2):
        sheet.cell(row=row, column=1, value=test)
    
    # Записываем заголовок "Исследование" в первую ячейку
    sheet.cell(row=1, column=1, value="Исследование").font = Font(bold=True)
    
    # Для каждой даты создаем три колонки
    for date in sorted(data_by_date.keys()):
        # Записываем дату как заголовок с "Результаты"
        date_str = date.strftime("%d.%m.%Y") + " Результаты"
        
        headers = [date_str, "Единицы", "Референс"]
        for i, header in enumerate(headers):
            cell = sheet.cell(row=1, column=current_col + i, value=header)
            cell.font = Font(bold=True)
        
        # Заполняем данные для каждого теста
        for row, test in enumerate(all_tests, start=2):
            # Ищем соответствующую запись для текущего теста
            test_data = next((data for data in data_by_date[date] if data[0] == test), None)
            if test_data:
                sheet.cell(row=row, column=current_col, value=test_data[1])  # Результат
                sheet.cell(row=row, column=current_col + 1, value=test_data[2])  # Единицы
                sheet.cell(row=row, column=current_col + 2, value=test_data[3])  # Референс
        
        current_col += 3
    
    # Автоподбор ширины колонок
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output_path)

@app.post("/upload/")
async def upload_pdf(files: list[UploadFile]):
    try:
        if not files:
            raise HTTPException(status_code=400, detail="Нет загруженных файлов.")

        excel_path = os.path.join(TEMP_DIR, "Medical_Analysis.xlsx")
        data_by_date = {}

        for file in files:
            try:
                date = extract_date_from_filename(file.filename)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=str(e))

            file_path = os.path.join(TEMP_DIR, file.filename)
            
            # Сохранение файла
            with open(file_path, "wb") as buffer:
                buffer.write(await file.read())

            # Обработка PDF
            raw_table_data = extract_data_from_pdf(file_path)
            
            # Фильтруем данные и добавляем их в словарь по датам
            filtered_data = [row for row in raw_table_data if is_relevant_row(row)]
            data_by_date[date] = filtered_data

            # Удаляем временный файл
            os.remove(file_path)

        # Генерация Excel с данными, организованными по датам
        create_excel_from_data(data_by_date, excel_path)

        return FileResponse(excel_path, filename="Medical_Analysis.xlsx")
    
    except Exception as e:
        print(f"Произошла ошибка: {e}")
        raise HTTPException(status_code=500, detail=f"Внутренняя ошибка сервера: {e}")
